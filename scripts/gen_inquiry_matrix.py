#!/usr/bin/env python3
"""
Generates docs/inquiry-matrix.xlsx with 6 sheets based on committed docs.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = "/workspaces/Patientenpfad/docs/inquiry-matrix.xlsx"

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
HDR_FILL  = PatternFill("solid", fgColor="1F497D")  # dark blue
HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL  = PatternFill("solid", fgColor="DCE6F1")  # light blue
NORM_FONT = Font(size=10)

def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")

def style_data_rows(ws, start=2):
    for i, row in enumerate(ws.iter_rows(min_row=start)):
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for cell in row:
            cell.font = NORM_FONT
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

def auto_width(ws, max_w=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_w)

def freeze_and_filter(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

# ---------------------------------------------------------------------------
# DATA
# ---------------------------------------------------------------------------

PROFILES = [
    # (id, label, decision_cp, intro_cp, shared_checkpoints, action_checkpoints, besonderheiten)
    (
        "ACUTE_CARE", "Akuttermin / offene Sprechstunde",
        "ACUTE_CARE_DECISION", "–",
        "INFECTIOUS_PROTOCOL",
        "ACUTE_OPEN_CONSULTATION_ACTION, ACUTE_BOOKING_INFO, CARE_CHANNEL_CHOICE, INFECTIOUS_DISEASE_NOTICE, INFECTIOUS_PROTOCOL",
        "Einziges Profil mit boundGlobalCheckpointIds; INFECTIOUS_PROTOCOL SHARED_BOTTOM"
    ),
    (
        "APPOINTMENT", "Termin",
        "–", "–",
        "–",
        "ACUTE_OPEN_CONSULTATION_ACTION, BOOK_APPOINTMENT, APPOINTMENT_BOOK_CHECKUP, APPOINTMENT_BOOK_FINDINGS_REVIEW, APPOINTMENT_BOOK_CHECKUP_SECOND, APPOINTMENT_BOOK_CHRONIC_CONTROL, APPOINTMENT_BOOK_EKG_ORDER, APPOINTMENT_INFO_CHECKUP, APPOINTMENT_INFO_CHRONIC_CONTROL, APPOINTMENT_INFO_FINDINGS_REVIEW, APPOINTMENT_INFO_EKG_ORDER, APPOINTMENT_INFO_CHECKUP_SECOND, APPOINTMENT_INFO_PSYCHOSOMATIC, APPOINTMENT_INFO_PHYSICAL_MEDICINE, APPOINTMENT_INFO_SURGERY, APPOINTMENT_INFO_SPECIALTY_VISIT, DOCUMENT_UPLOAD, INSURANCE_DATA_APP_TRANSFER",
        "Kein Decision-CP; größtes Profil (18 bound actions); APPOINTMENT_TYPE_QUESTION schaltet 9 Info-Actions frei"
    ),
    (
        "AU", "AU / Arbeitsunfähigkeitsbescheinigung",
        "AU_DECISION", "–",
        "–",
        "AU_NEW_PATIENT_3DAY_LIMIT, DIGITAL_REQUEST, DIGITAL_REQUEST_PROCESSING_TIME, ACUTE_OPEN_CONSULTATION_ACTION, CARE_CHANNEL_CHOICE, CONTROL_APPOINTMENT_RECOMMENDED, INSURANCE_DATA_APP_TRANSFER",
        "DIGITAL_REQUEST_PROCESSING_TIME config-abhängig; CONTROL_APPOINTMENT_RECOMMENDED SHARED_BOTTOM"
    ),
    (
        "PRESCRIPTION", "Rezept",
        "PRESCRIPTION_DECISION", "–",
        "–",
        "E_RECIPE_USE, PHARMACY_INFORMATION, INSURANCE_DATA_APP_TRANSFER",
        "10 actionGuidanceRules; umfangreichstes GuidanceRule-Set"
    ),
    (
        "MEDICAL_DOCUMENTS", "Atteste / Bescheinigungen",
        "MEDICAL_DOCUMENTS_DECISION", "–",
        "–",
        "DIGITAL_REQUEST, BOOK_APPOINTMENT, DOCUMENT_UPLOAD, PAYMENT_ONSITE_INFO",
        "availableActionIds leer"
    ),
    (
        "HEILMITTELVERORDNUNG", "Heilmittelverordnung",
        "–", "–",
        "–",
        "DIGITAL_REQUEST, DOCUMENT_UPLOAD, BOOK_APPOINTMENT",
        "Keine availableActionIds; alle Actions strikt konditioniert"
    ),
    (
        "REFERRAL", "Überweisung",
        "REFERRAL_DECISION", "–",
        "–",
        "REF_BOOKING_CODE_PROCESS, REF_ORIGINAL_VS_PDF, INSURANCE_DATA_APP_TRANSFER",
        "REF_ORIGINAL_VS_PDF immer aktiv"
    ),
    (
        "HOSPITAL_ADMISSION", "Krankenhauseinweisung",
        "HOSPITAL_ADMISSION_DECISION", "–",
        "–",
        "CONTROL_APPOINTMENT_RECOMMENDED, TRANSPORT_QUESTIONNAIRE_REQUEST",
        "Beide Actions immer aktiv"
    ),
    (
        "IMMUNIZATION", "Impfung",
        "IMMUNIZATION_DECISION", "–",
        "–",
        "IMMUNIZATION_BOOK_VACCINATION, IMMUNIZATION_BOOK_COUNSELING, IMMUNIZATION_BRING_VACCINATION_RECORD",
        "BRING_RECORD standardmäßig aktiv, suppressed bei TRAVEL_MEDICINE"
    ),
    (
        "LAB", "Labor",
        "LAB_DECISION", "–",
        "BILLING_COST_NOT_COVERED, APPOINTMENT_DATA_INCOMPLETE",
        "LAB_APPOINTMENT_INTERNAL, LAB_APPOINTMENT_EXTERNAL, LAB_BRING_REFERRAL, LAB_COST_COVERED_BY_REFERRAL, LAB_SELF_PAYER_NOTE, LAB_FASTING_REQUIRED, LAB_RESULT_TIME, LAB_SAMPLE_FOLLOWUP_APPOINTMENT_RECOMMENDED",
        "LAB_APPOINTMENT_INTERNAL config-abhängig; LAB_FASTING suppresst bei MPU"
    ),
    (
        "SAMPLE_COLLECTION", "Urin- und Stuhlprobe",
        "SAMPLE_COLLECTION_DECISION", "–",
        "–",
        "URINE_SAMPLE_INSTRUCTIONS, STOOL_SAMPLE_INSTRUCTIONS, SAMPLE_HANDOVER, URINE_SAMPLE_ONSITE, LAB_SAMPLE_FOLLOWUP_APPOINTMENT_RECOMMENDED",
        "Kleinstes Profil (1 specific CP)"
    ),
    (
        "ONBOARDING", "Patientenaufnahme / Registrierung",
        "–", "–",
        "–",
        "ONBOARDING_IDENTITY_CLARIFICATION_REQUIRED, ONBOARDING_PROVIDE_IDENTITY_DATA, ONBOARDING_DATA_MISSING_CONTEXT, ONBOARDING_WRONG_PRACTICE_NOTICE, DOCUMENT_UPLOAD, INSURANCE_DATA_APP_TRANSFER",
        "ONBOARDING_DOCTOLIB_INFO config-abhängig"
    ),
    (
        "BILLING", "Abrechnung",
        "–", "–",
        "BILLING_COST_NOT_COVERED",
        "BILLING_NOT_COVERED_BY_STATUTORY, BILLING_GOA_BILLING, BILLING_ONSITE_PAYMENT, BILLING_CONTACT_EXTERNAL_PARTY, BILLING_ADDRESS_UPDATE_REQUESTED, INSURANCE_DATA_APP_TRANSFER",
        "BILLING_INVOICE_TIMING config-abhängig; BILLING_COST_NOT_COVERED auch in LAB"
    ),
    (
        "TECH_SUPPORT", "Technische Probleme / Digitale Infrastruktur",
        "–", "–",
        "–",
        "– (keine bound actions)",
        "TECH_VIDEO_NOT_WORKING config-abhängig; keine bound action-checkpoints"
    ),
]

# ---------------------------------------------------------------------------
# Sheet 1: Profiles
# ---------------------------------------------------------------------------
def build_profiles(wb):
    ws = wb.create_sheet("Profiles")
    headers = [
        "Profile ID", "Label", "Decision Checkpoint", "Intro Checkpoint",
        "Shared Checkpoints", "Action Checkpoints", "Besonderheiten"
    ]
    ws.append(headers)
    for p in PROFILES:
        ws.append(list(p))
    style_header_row(ws)
    style_data_rows(ws)
    auto_width(ws)
    freeze_and_filter(ws)

# ---------------------------------------------------------------------------
# Checkpoint data
# ---------------------------------------------------------------------------

# (id, typ, placement, audience, kurzbeschreibung, shared_global, config_abhaengig, config_felder, deaktivierbar, hinweise)
CHECKPOINTS = [
    # --- Decision ---
    ("AU_DECISION","DECISION","ATTACHED","OFFICE","Weichenstellung AU-Typ","Nein","Nein","","Nein",""),
    ("PRESCRIPTION_DECISION","DECISION","ATTACHED","OFFICE","Weichenstellung Rezept-Typ","Nein","Nein","","Nein",""),
    ("LAB_DECISION","DECISION","ATTACHED","OFFICE","Weichenstellung Labor-Typ","Nein","Nein","","Nein",""),
    ("SAMPLE_COLLECTION_DECISION","DECISION","ATTACHED","OFFICE","Weichenstellung Probenart","Nein","Nein","","Nein",""),
    ("ACUTE_CARE_DECISION","DECISION","ATTACHED","OFFICE","Weichenstellung Akut-Typ","Nein","Nein","","Nein",""),
    ("REFERRAL_DECISION","DECISION","ATTACHED","OFFICE","Weichenstellung Überweisungs-Typ","Nein","Nein","","Nein",""),
    ("IMMUNIZATION_DECISION","DECISION","ATTACHED","OFFICE","Weichenstellung Impf-Typ","Nein","Nein","","Nein",""),
    ("MEDICAL_DOCUMENTS_DECISION","DECISION","ATTACHED","OFFICE","Weichenstellung Dokumentenart","Nein","Nein","","Nein",""),
    ("HOSPITAL_ADMISSION_DECISION","DECISION","ATTACHED","OFFICE","Weichenstellung Einweisungsart","Nein","Nein","","Nein",""),
    # --- Explanation SPECIFIC ---
    ("ACUTE_CARE_GENERAL_INFO","EXPLANATION","ATTACHED","PATIENT","Allgemeine Info Akuttermin","Nein","Nein","","Ja",""),
    ("ACUTE_BOOKING_INFO","EXPLANATION","ATTACHED","PATIENT","Buchungsinfo Akuttermin","Nein","Nein","","Ja",""),
    ("INFECTIOUS_DISEASE_NOTICE","EXPLANATION","ATTACHED","PATIENT","Hinweis Infektionskrankheit","Nein","Nein","","Ja",""),
    ("AU_NEW_PATIENT_3DAY_LIMIT","EXPLANATION","ATTACHED","PATIENT","3-Tage-Limit Neupatient AU","Nein","Nein","","Ja",""),
    ("PRESCRIPTION_GENERAL_INFO","EXPLANATION","ATTACHED","PATIENT","Allgemeine Info Rezept","Nein","Nein","","Ja",""),
    ("PRESCRIPTION_SPECIALIST_REQUIRED","EXPLANATION","ATTACHED","PATIENT","Facharztvorbehalt Rezept","Nein","Nein","","Ja",""),
    ("PRESCRIPTION_NARCOTIC_NOT_POSSIBLE","EXPLANATION","ATTACHED","PATIENT","BTM-Rezept nicht möglich","Nein","Nein","","Ja",""),
    ("PRESCRIPTION_NEW_PATIENT","EXPLANATION","ATTACHED","PATIENT","Rezept Neupatient","Nein","Nein","","Ja",""),
    ("PRESCRIPTION_FOLLOW_UP_APPOINTMENT","EXPLANATION","ATTACHED","PATIENT","Folgetermin für Rezept nötig","Nein","Nein","","Ja",""),
    ("PRESCRIPTION_MISSING_DIAGNOSIS","EXPLANATION","ATTACHED","PATIENT","Fehlende Diagnose","Nein","Nein","","Ja",""),
    ("MEDICAL_DOCUMENTS_GENERAL_INFO","EXPLANATION","ATTACHED","PATIENT","Allgemeine Info Atteste","Nein","Nein","","Ja",""),
    ("MEDICAL_DOCUMENTS_APPOINTMENT_REQUIRED","EXPLANATION","ATTACHED","PATIENT","Termin für Attest erforderlich","Nein","Nein","","Ja",""),
    ("MEDICAL_DOCUMENTS_NOT_POSSIBLE","EXPLANATION","ATTACHED","PATIENT","Dokument nicht ausstellbar","Nein","Nein","","Ja",""),
    ("HEILMITTELVERORDNUNG_GENERAL_INFO","EXPLANATION","ATTACHED","PATIENT","Allgemeine Info HMV","Nein","Nein","","Ja",""),
    ("REFERRAL_GENERAL_INFO","EXPLANATION","ATTACHED","PATIENT","Allgemeine Info Überweisung","Nein","Nein","","Ja",""),
    ("REFERRAL_NOT_POSSIBLE","EXPLANATION","ATTACHED","PATIENT","Überweisung nicht möglich","Nein","Nein","","Ja",""),
    ("REFERRAL_APPOINTMENT_REQUIRED","EXPLANATION","ATTACHED","PATIENT","Termin für Überweisung nötig","Nein","Nein","","Ja",""),
    ("HOSPITAL_ADMISSION_GENERAL_INFO","EXPLANATION","ATTACHED","PATIENT","Allgemeine Info Einweisung","Nein","Nein","","Ja",""),
    ("IMMUNIZATION_GENERAL_INFO","EXPLANATION","ATTACHED","PATIENT","Allgemeine Info Impfung","Nein","Nein","","Ja",""),
    ("IMMUNIZATION_NOT_AVAILABLE","EXPLANATION","ATTACHED","PATIENT","Impfstoff nicht verfügbar","Nein","Nein","","Ja",""),
    ("IMMUNIZATION_APPOINTMENT_REQUIRED","EXPLANATION","ATTACHED","PATIENT","Termin für Impfberatung","Nein","Nein","","Ja",""),
    ("LAB_GENERAL_INFO","EXPLANATION","ATTACHED","PATIENT","Allgemeine Info Labor","Nein","Nein","","Ja",""),
    ("LAB_EXTERNAL_LAB_INFO","EXPLANATION","ATTACHED","PATIENT","Info externes Labor","Nein","Nein","","Ja",""),
    ("SAMPLE_COLLECTION_GENERAL_INFO","EXPLANATION","ATTACHED","PATIENT","Allgemeine Info Probenabgabe","Nein","Nein","","Ja",""),
    ("ONBOARDING_GENERAL_INFO","EXPLANATION","ATTACHED","PATIENT","Allgemeine Info Aufnahme","Nein","Nein","","Ja",""),
    ("BILLING_GENERAL_INFO","EXPLANATION","ATTACHED","PATIENT","Allgemeine Info Abrechnung","Nein","Nein","","Ja",""),
    ("TECH_SUPPORT_GENERAL_INFO","EXPLANATION","ATTACHED","PATIENT","Allgemeine Info Tech-Support","Nein","Nein","","Ja",""),
    # --- Global MODULAR Explanation ---
    ("INFECTIOUS_PROTOCOL","EXPLANATION","SHARED_BOTTOM","PATIENT","Infektionsschutz-Protokoll","Ja","Nein","","Ja","Bound global in ACUTE_CARE"),
    ("TRANSPORT_QUESTIONNAIRE_REQUEST","EXPLANATION","ATTACHED","PATIENT","Transportmittel-Fragebogen","Ja","Nein","","Ja",""),
    ("REQUIRED_INFORMATION_COMPLETE","EXPLANATION","ATTACHED","PATIENT","Informationen vollständig","Ja","Nein","","Ja",""),
    ("CONTACT_PERSON_INFO","EXPLANATION","SHARED_BOTTOM","PATIENT","Ansprechperson Praxis","Ja","Nein","","Ja","Verwendet in AU, REF, RX, ONB"),
    ("PAYMENT_ONSITE_INFO","EXPLANATION","ATTACHED","PATIENT","Zahlung vor Ort","Ja","Nein","","Ja",""),
    ("APPOINTMENT_DATA_INCOMPLETE","EXPLANATION","ATTACHED","PATIENT","Termindaten unvollständig","Ja","Nein","","Ja","Auch in LAB"),
    ("BILLING_COST_NOT_COVERED","EXPLANATION","ATTACHED","PATIENT","Kosten nicht gedeckt","Ja","Nein","","Nein","In LAB und BIL"),
    ("APPOINTMENT_INFO_GENERAL","EXPLANATION","ATTACHED","PATIENT","Allgemeine Termininfo","Ja","Nein","","Ja",""),
    ("ONBOARDING_DOCTOLIB_INFO","EXPLANATION","ATTACHED","PATIENT","Doctolib-Plattforminfo","Nein","Ja","uploadPlatformName, uploadPlatformAccountLabel","Ja","Nur ONB"),
    ("BILLING_INVOICE_TIMING","EXPLANATION","ATTACHED","PATIENT","Abrechnungszyklus-Info","Nein","Ja","billingCycleLabel","Ja","Nur BIL"),
    ("TECH_VIDEO_NOT_WORKING","EXPLANATION","ATTACHED","PATIENT","Video-Support-Info","Nein","Ja","videoSupportContact","Ja","Nur TECH"),
    # --- Global Action SHARED_BOTTOM ---
    ("DIGITAL_REQUEST","ACTION","SHARED_BOTTOM","PATIENT","Digitale Anfrage stellen","Ja","Nein","","Nein","In fast allen Profilen"),
    ("BOOK_APPOINTMENT","ACTION","SHARED_BOTTOM","PATIENT","Termin buchen","Ja","Nein","","Nein","In fast allen Profilen"),
    ("INSURANCE_DATA_APP_TRANSFER","ACTION","SHARED_BOTTOM","PATIENT","Versicherungsdaten übertragen","Ja","Nein","","Ja","AU, RX, REF, APP, ONB, BIL"),
    ("CONTROL_APPOINTMENT_RECOMMENDED","ACTION","SHARED_BOTTOM","PATIENT","Kontrolltermin empfohlen","Ja","Nein","","Ja","AU, HOSP, RX"),
    # --- Global Action ATTACHED ---
    ("ACUTE_OPEN_CONSULTATION_ACTION","ACTION","ATTACHED","PATIENT","Offene Sprechstunde Buchung","Ja","Ja","openConsultationDays, openConsultationHours, openConsultationCapacityLimited","Ja","AU, ACU, APP"),
    ("CARE_CHANNEL_CHOICE","ACTION","ATTACHED","PATIENT","Kanal-Auswahl","Ja","Nein","","Ja","AU, ACU"),
    # --- Specific Actions APPOINTMENT ---
    ("APPOINTMENT_BOOK_CHECKUP","ACTION","ATTACHED","PATIENT","Check-up Termin buchen","Nein","Nein","","Ja",""),
    ("APPOINTMENT_BOOK_FINDINGS_REVIEW","ACTION","ATTACHED","PATIENT","Befundbesprechung buchen","Nein","Ja","bookingCalendarName, findingsReviewBookingCode","Ja",""),
    ("APPOINTMENT_BOOK_CHECKUP_SECOND","ACTION","ATTACHED","PATIENT","Check-up 2. Termin buchen","Nein","Ja","bookingCalendarName, checkupSecondBookingCode","Ja",""),
    ("APPOINTMENT_BOOK_CHRONIC_CONTROL","ACTION","ATTACHED","PATIENT","Chronikerkontrolle buchen","Nein","Ja","bookingCalendarName, chronicControlBookingCode","Ja",""),
    ("APPOINTMENT_BOOK_EKG_ORDER","ACTION","ATTACHED","PATIENT","EKG-Termin buchen","Nein","Ja","bookingCalendarName, doctorOrderBookingCode","Ja",""),
    ("APPOINTMENT_INFO_CHECKUP","ACTION","ATTACHED","PATIENT","Info Check-up","Nein","Nein","","Ja",""),
    ("APPOINTMENT_INFO_CHRONIC_CONTROL","ACTION","ATTACHED","PATIENT","Info Chronikertermin","Nein","Nein","","Ja",""),
    ("APPOINTMENT_INFO_FINDINGS_REVIEW","ACTION","ATTACHED","PATIENT","Info Befundbesprechung","Nein","Nein","","Ja",""),
    ("APPOINTMENT_INFO_EKG_ORDER","ACTION","ATTACHED","PATIENT","Info EKG-Termin","Nein","Nein","","Ja",""),
    ("APPOINTMENT_INFO_CHECKUP_SECOND","ACTION","ATTACHED","PATIENT","Info Check-up 2. Termin","Nein","Nein","","Ja",""),
    ("APPOINTMENT_INFO_PSYCHOSOMATIC","ACTION","ATTACHED","PATIENT","Info Psychosomatik","Nein","Nein","","Ja",""),
    ("APPOINTMENT_INFO_PHYSICAL_MEDICINE","ACTION","ATTACHED","PATIENT","Info Physikalische Medizin","Nein","Nein","","Ja",""),
    ("APPOINTMENT_INFO_SURGERY","ACTION","ATTACHED","PATIENT","Info Chirurgie","Nein","Nein","","Ja",""),
    ("APPOINTMENT_INFO_SPECIALTY_VISIT","ACTION","ATTACHED","PATIENT","Info Spezialsprechstunde","Nein","Nein","","Ja",""),
    # --- Specific Actions AU ---
    ("DIGITAL_REQUEST_PROCESSING_TIME","ACTION","ATTACHED","PATIENT","Bearbeitungszeit digitale Anfrage","Nein","Ja","digitalRequestProcessingTimeMin, digitalRequestProcessingTimeMax, digitalRequestProcessingTimeUnit","Ja","Nur AU"),
    # --- Specific Actions PRESCRIPTION ---
    ("E_RECIPE_USE","ACTION","ATTACHED","PATIENT","E-Rezept verwenden","Nein","Nein","","Ja",""),
    ("PHARMACY_INFORMATION","ACTION","ATTACHED","PATIENT","Apothekeninfo","Nein","Nein","","Ja",""),
    # --- Specific Actions MEDICAL_DOCUMENTS ---
    # (DIGITAL_REQUEST, BOOK_APPOINTMENT, DOCUMENT_UPLOAD, PAYMENT_ONSITE_INFO already listed)
    # --- Specific Actions REFERRAL ---
    ("REF_BOOKING_CODE_PROCESS","ACTION","ATTACHED","PATIENT","Buchungscode-Prozess Überweisung","Nein","Nein","","Ja",""),
    ("REF_ORIGINAL_VS_PDF","ACTION","ATTACHED","PATIENT","Original vs. PDF Überweisung","Nein","Nein","","Nein","Immer aktiv"),
    # --- Specific Actions HOSPITAL_ADMISSION ---
    # (CONTROL_APPOINTMENT_RECOMMENDED, TRANSPORT_QUESTIONNAIRE_REQUEST already listed)
    # --- Specific Actions IMMUNIZATION ---
    ("IMMUNIZATION_BOOK_VACCINATION","ACTION","ATTACHED","PATIENT","Impftermin buchen","Nein","Nein","","Ja",""),
    ("IMMUNIZATION_BOOK_COUNSELING","ACTION","ATTACHED","PATIENT","Impfberatungstermin buchen","Nein","Nein","","Ja",""),
    ("IMMUNIZATION_BRING_VACCINATION_RECORD","ACTION","ATTACHED","PATIENT","Impfausweis mitbringen","Nein","Nein","","Ja","Suppressed bei TRAVEL_MEDICINE"),
    # --- Specific Actions LAB ---
    ("LAB_APPOINTMENT_INTERNAL","ACTION","ATTACHED","PATIENT","Interner Labortermin","Nein","Ja","doctorOrderBookingCode","Ja",""),
    ("LAB_APPOINTMENT_EXTERNAL","ACTION","ATTACHED","PATIENT","Externer Labortermin","Nein","Nein","","Ja",""),
    ("LAB_BRING_REFERRAL","ACTION","ATTACHED","PATIENT","Überweisungsschein mitbringen","Nein","Nein","","Ja",""),
    ("LAB_COST_COVERED_BY_REFERRAL","ACTION","ATTACHED","PATIENT","Kosten per Überweisung gedeckt","Nein","Nein","","Ja",""),
    ("LAB_SELF_PAYER_NOTE","ACTION","ATTACHED","PATIENT","Selbstzahler-Hinweis","Nein","Nein","","Ja",""),
    ("LAB_FASTING_REQUIRED","ACTION","ATTACHED","PATIENT","Nüchternheit erforderlich","Nein","Nein","","Ja","Suppressed bei MPU"),
    ("LAB_RESULT_TIME","ACTION","ATTACHED","PATIENT","Ergebniszeit Labor","Nein","Nein","","Ja","Auch in SMP"),
    ("LAB_SAMPLE_FOLLOWUP_APPOINTMENT_RECOMMENDED","ACTION","ATTACHED","PATIENT","Folgetermin Probe","Nein","Nein","","Ja","Auch in SMP"),
    # --- Specific Actions SAMPLE_COLLECTION ---
    ("URINE_SAMPLE_INSTRUCTIONS","ACTION","ATTACHED","PATIENT","Anleitung Urinprobe","Nein","Nein","","Ja",""),
    ("STOOL_SAMPLE_INSTRUCTIONS","ACTION","ATTACHED","PATIENT","Anleitung Stuhlprobe","Nein","Nein","","Ja",""),
    ("SAMPLE_HANDOVER","ACTION","ATTACHED","PATIENT","Probenabgabe Info","Nein","Nein","","Ja",""),
    ("URINE_SAMPLE_ONSITE","ACTION","ATTACHED","PATIENT","Urinprobe vor Ort","Nein","Nein","","Ja",""),
    # --- Specific Actions ONBOARDING ---
    ("ONBOARDING_IDENTITY_CLARIFICATION_REQUIRED","ACTION","ATTACHED","PATIENT","Identitätsklärung nötig","Nein","Nein","","Nein",""),
    ("ONBOARDING_PROVIDE_IDENTITY_DATA","ACTION","ATTACHED","PATIENT","Identitätsdaten übermitteln","Nein","Nein","","Ja",""),
    ("ONBOARDING_DATA_MISSING_CONTEXT","ACTION","ATTACHED","PATIENT","Fehlende Daten Kontext","Nein","Nein","","Ja",""),
    ("ONBOARDING_WRONG_PRACTICE_NOTICE","ACTION","ATTACHED","PATIENT","Falsche Praxis Hinweis","Nein","Nein","","Ja",""),
    # --- Specific Actions BILLING ---
    ("BILLING_NOT_COVERED_BY_STATUTORY","ACTION","ATTACHED","PATIENT","Nicht von GKV gedeckt","Nein","Nein","","Ja",""),
    ("BILLING_GOA_BILLING","ACTION","ATTACHED","PATIENT","GOÄ-Abrechnung","Nein","Nein","","Ja",""),
    ("BILLING_ONSITE_PAYMENT","ACTION","ATTACHED","PATIENT","Zahlung vor Ort","Nein","Nein","","Ja",""),
    ("BILLING_CONTACT_EXTERNAL_PARTY","ACTION","ATTACHED","PATIENT","Externer Ansprechpartner","Nein","Nein","","Ja",""),
    ("BILLING_ADDRESS_UPDATE_REQUESTED","ACTION","ATTACHED","PATIENT","Adressänderung angefordert","Nein","Nein","","Ja",""),
    # --- DOCUMENT_UPLOAD (config-abhängig, mehrere Profile) ---
    ("DOCUMENT_UPLOAD","ACTION","ATTACHED","PATIENT","Dokument hochladen","Ja","Ja","uploadPlatformName, uploadPlatformAccountLabel","Ja","RX, HMV, HOSP, APP, DOC, ONB"),
    # --- Intro/SectionIntro ---
    ("MESSAGE_INTRO_ACUTE_CARE","INTRO","ATTACHED","PATIENT","Einleitung Akuttermin","Nein","Nein","","Ja",""),
    ("MESSAGE_INTRO_APPOINTMENT","INTRO","ATTACHED","PATIENT","Einleitung Termin","Nein","Nein","","Ja",""),
    ("MESSAGE_INTRO_AU","INTRO","ATTACHED","PATIENT","Einleitung AU","Nein","Nein","","Ja",""),
    ("MESSAGE_INTRO_PRESCRIPTION","INTRO","ATTACHED","PATIENT","Einleitung Rezept","Nein","Nein","","Ja",""),
    ("MESSAGE_INTRO_MEDICAL_DOCUMENTS","INTRO","ATTACHED","PATIENT","Einleitung Atteste","Nein","Nein","","Ja",""),
    ("MESSAGE_INTRO_HEILMITTELVERORDNUNG","INTRO","ATTACHED","PATIENT","Einleitung HMV","Nein","Nein","","Ja",""),
    ("MESSAGE_INTRO_REFERRAL","INTRO","ATTACHED","PATIENT","Einleitung Überweisung","Nein","Nein","","Ja",""),
    ("MESSAGE_INTRO_HOSPITAL_ADMISSION","INTRO","ATTACHED","PATIENT","Einleitung Einweisung","Nein","Nein","","Ja",""),
    ("MESSAGE_INTRO_IMMUNIZATION","INTRO","ATTACHED","PATIENT","Einleitung Impfung","Nein","Nein","","Ja",""),
    ("MESSAGE_INTRO_LAB","INTRO","ATTACHED","PATIENT","Einleitung Labor","Nein","Nein","","Ja",""),
    ("MESSAGE_INTRO_SAMPLE_COLLECTION","INTRO","ATTACHED","PATIENT","Einleitung Probenabgabe","Nein","Nein","","Ja",""),
    ("MESSAGE_INTRO_ONBOARDING","INTRO","ATTACHED","PATIENT","Einleitung Aufnahme","Nein","Nein","","Ja",""),
    ("MESSAGE_INTRO_BILLING","INTRO","ATTACHED","PATIENT","Einleitung Abrechnung","Nein","Nein","","Ja",""),
    ("MESSAGE_INTRO_TECH_SUPPORT","INTRO","ATTACHED","PATIENT","Einleitung Tech-Support","Nein","Nein","","Ja",""),
]

# ---------------------------------------------------------------------------
# Sheet 2: Checkpoints
# ---------------------------------------------------------------------------
def build_checkpoints(wb):
    ws = wb.create_sheet("Checkpoints")
    headers = [
        "Checkpoint ID","Typ","Placement","Audience","Kurzbeschreibung",
        "Shared/global","Config-abhängig","Config-Felder","Sicher deaktivierbar?","Hinweise"
    ]
    ws.append(headers)
    for cp in CHECKPOINTS:
        ws.append(list(cp))
    style_header_row(ws)
    style_data_rows(ws)
    auto_width(ws)
    freeze_and_filter(ws)

# ---------------------------------------------------------------------------
# Profile-Checkpoint matrix data
# ---------------------------------------------------------------------------
# Each row: (profile_id, checkpoint_id, typ, placement, trigger_bedingung, shared, is_action, is_decision, is_intro)
MATRIX_ROWS = []

def add_row(profile, cp_id, typ, placement, trigger, shared, action, decision, intro):
    MATRIX_ROWS.append((profile, cp_id, typ, placement, trigger, shared, action, decision, intro))

# ACUTE_CARE
add_row("ACUTE_CARE","ACUTE_CARE_DECISION","DECISION","ATTACHED","immer","Nein","Nein","Ja","Nein")
add_row("ACUTE_CARE","ACUTE_OPEN_CONSULTATION_ACTION","ACTION","ATTACHED","config-abhängig","Ja","Ja","Nein","Nein")
add_row("ACUTE_CARE","ACUTE_BOOKING_INFO","EXPLANATION","ATTACHED","bedingt","Nein","Nein","Nein","Nein")
add_row("ACUTE_CARE","CARE_CHANNEL_CHOICE","ACTION","ATTACHED","bedingt","Ja","Ja","Nein","Nein")
add_row("ACUTE_CARE","INFECTIOUS_DISEASE_NOTICE","EXPLANATION","ATTACHED","bedingt","Nein","Nein","Nein","Nein")
add_row("ACUTE_CARE","INFECTIOUS_PROTOCOL","EXPLANATION","SHARED_BOTTOM","bedingt","Ja","Nein","Nein","Nein")
# APPOINTMENT
add_row("APPOINTMENT","ACUTE_OPEN_CONSULTATION_ACTION","ACTION","ATTACHED","config-abhängig","Ja","Ja","Nein","Nein")
add_row("APPOINTMENT","BOOK_APPOINTMENT","ACTION","SHARED_BOTTOM","bedingt","Ja","Ja","Nein","Nein")
add_row("APPOINTMENT","APPOINTMENT_BOOK_CHECKUP","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("APPOINTMENT","APPOINTMENT_BOOK_FINDINGS_REVIEW","ACTION","ATTACHED","config-abhängig","Nein","Ja","Nein","Nein")
add_row("APPOINTMENT","APPOINTMENT_BOOK_CHECKUP_SECOND","ACTION","ATTACHED","config-abhängig","Nein","Ja","Nein","Nein")
add_row("APPOINTMENT","APPOINTMENT_BOOK_CHRONIC_CONTROL","ACTION","ATTACHED","config-abhängig","Nein","Ja","Nein","Nein")
add_row("APPOINTMENT","APPOINTMENT_BOOK_EKG_ORDER","ACTION","ATTACHED","config-abhängig","Nein","Ja","Nein","Nein")
add_row("APPOINTMENT","APPOINTMENT_INFO_CHECKUP","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("APPOINTMENT","APPOINTMENT_INFO_CHRONIC_CONTROL","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("APPOINTMENT","APPOINTMENT_INFO_FINDINGS_REVIEW","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("APPOINTMENT","APPOINTMENT_INFO_EKG_ORDER","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("APPOINTMENT","APPOINTMENT_INFO_CHECKUP_SECOND","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("APPOINTMENT","APPOINTMENT_INFO_PSYCHOSOMATIC","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("APPOINTMENT","APPOINTMENT_INFO_PHYSICAL_MEDICINE","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("APPOINTMENT","APPOINTMENT_INFO_SURGERY","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("APPOINTMENT","APPOINTMENT_INFO_SPECIALTY_VISIT","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("APPOINTMENT","DOCUMENT_UPLOAD","ACTION","ATTACHED","config-abhängig","Ja","Ja","Nein","Nein")
add_row("APPOINTMENT","INSURANCE_DATA_APP_TRANSFER","ACTION","SHARED_BOTTOM","bedingt","Ja","Ja","Nein","Nein")
# AU
add_row("AU","AU_DECISION","DECISION","ATTACHED","immer","Nein","Nein","Ja","Nein")
add_row("AU","AU_NEW_PATIENT_3DAY_LIMIT","EXPLANATION","ATTACHED","bedingt","Nein","Nein","Nein","Nein")
add_row("AU","DIGITAL_REQUEST","ACTION","SHARED_BOTTOM","bedingt","Ja","Ja","Nein","Nein")
add_row("AU","DIGITAL_REQUEST_PROCESSING_TIME","ACTION","ATTACHED","config-abhängig","Nein","Ja","Nein","Nein")
add_row("AU","ACUTE_OPEN_CONSULTATION_ACTION","ACTION","ATTACHED","config-abhängig","Ja","Ja","Nein","Nein")
add_row("AU","CARE_CHANNEL_CHOICE","ACTION","ATTACHED","bedingt","Ja","Ja","Nein","Nein")
add_row("AU","CONTROL_APPOINTMENT_RECOMMENDED","ACTION","SHARED_BOTTOM","bedingt","Ja","Ja","Nein","Nein")
add_row("AU","INSURANCE_DATA_APP_TRANSFER","ACTION","SHARED_BOTTOM","bedingt","Ja","Ja","Nein","Nein")
# PRESCRIPTION
add_row("PRESCRIPTION","PRESCRIPTION_DECISION","DECISION","ATTACHED","immer","Nein","Nein","Ja","Nein")
add_row("PRESCRIPTION","E_RECIPE_USE","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("PRESCRIPTION","PHARMACY_INFORMATION","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("PRESCRIPTION","INSURANCE_DATA_APP_TRANSFER","ACTION","SHARED_BOTTOM","bedingt","Ja","Ja","Nein","Nein")
add_row("PRESCRIPTION","CONTROL_APPOINTMENT_RECOMMENDED","ACTION","SHARED_BOTTOM","bedingt","Ja","Ja","Nein","Nein")
# MEDICAL_DOCUMENTS
add_row("MEDICAL_DOCUMENTS","MEDICAL_DOCUMENTS_DECISION","DECISION","ATTACHED","immer","Nein","Nein","Ja","Nein")
add_row("MEDICAL_DOCUMENTS","DIGITAL_REQUEST","ACTION","SHARED_BOTTOM","bedingt","Ja","Ja","Nein","Nein")
add_row("MEDICAL_DOCUMENTS","BOOK_APPOINTMENT","ACTION","SHARED_BOTTOM","bedingt","Ja","Ja","Nein","Nein")
add_row("MEDICAL_DOCUMENTS","DOCUMENT_UPLOAD","ACTION","ATTACHED","config-abhängig","Ja","Ja","Nein","Nein")
add_row("MEDICAL_DOCUMENTS","PAYMENT_ONSITE_INFO","EXPLANATION","ATTACHED","bedingt","Ja","Nein","Nein","Nein")
# HEILMITTELVERORDNUNG
add_row("HEILMITTELVERORDNUNG","DIGITAL_REQUEST","ACTION","SHARED_BOTTOM","bedingt","Ja","Ja","Nein","Nein")
add_row("HEILMITTELVERORDNUNG","DOCUMENT_UPLOAD","ACTION","ATTACHED","config-abhängig","Ja","Ja","Nein","Nein")
add_row("HEILMITTELVERORDNUNG","BOOK_APPOINTMENT","ACTION","SHARED_BOTTOM","bedingt","Ja","Ja","Nein","Nein")
# REFERRAL
add_row("REFERRAL","REFERRAL_DECISION","DECISION","ATTACHED","immer","Nein","Nein","Ja","Nein")
add_row("REFERRAL","REF_BOOKING_CODE_PROCESS","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("REFERRAL","REF_ORIGINAL_VS_PDF","ACTION","ATTACHED","immer aktiv","Nein","Ja","Nein","Nein")
add_row("REFERRAL","INSURANCE_DATA_APP_TRANSFER","ACTION","SHARED_BOTTOM","bedingt","Ja","Ja","Nein","Nein")
# HOSPITAL_ADMISSION
add_row("HOSPITAL_ADMISSION","HOSPITAL_ADMISSION_DECISION","DECISION","ATTACHED","immer","Nein","Nein","Ja","Nein")
add_row("HOSPITAL_ADMISSION","CONTROL_APPOINTMENT_RECOMMENDED","ACTION","SHARED_BOTTOM","immer aktiv","Ja","Ja","Nein","Nein")
add_row("HOSPITAL_ADMISSION","TRANSPORT_QUESTIONNAIRE_REQUEST","EXPLANATION","ATTACHED","immer aktiv","Ja","Nein","Nein","Nein")
# IMMUNIZATION
add_row("IMMUNIZATION","IMMUNIZATION_DECISION","DECISION","ATTACHED","immer","Nein","Nein","Ja","Nein")
add_row("IMMUNIZATION","IMMUNIZATION_BOOK_VACCINATION","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("IMMUNIZATION","IMMUNIZATION_BOOK_COUNSELING","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("IMMUNIZATION","IMMUNIZATION_BRING_VACCINATION_RECORD","ACTION","ATTACHED","standard aktiv; suppressed bei TRAVEL_MEDICINE","Nein","Ja","Nein","Nein")
# LAB
add_row("LAB","LAB_DECISION","DECISION","ATTACHED","immer","Nein","Nein","Ja","Nein")
add_row("LAB","BILLING_COST_NOT_COVERED","EXPLANATION","ATTACHED","bedingt","Ja","Nein","Nein","Nein")
add_row("LAB","APPOINTMENT_DATA_INCOMPLETE","EXPLANATION","ATTACHED","bedingt","Ja","Nein","Nein","Nein")
add_row("LAB","LAB_APPOINTMENT_INTERNAL","ACTION","ATTACHED","config-abhängig","Nein","Ja","Nein","Nein")
add_row("LAB","LAB_APPOINTMENT_EXTERNAL","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("LAB","LAB_BRING_REFERRAL","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("LAB","LAB_COST_COVERED_BY_REFERRAL","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("LAB","LAB_SELF_PAYER_NOTE","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("LAB","LAB_FASTING_REQUIRED","ACTION","ATTACHED","bedingt; suppressed bei MPU","Nein","Ja","Nein","Nein")
add_row("LAB","LAB_RESULT_TIME","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("LAB","LAB_SAMPLE_FOLLOWUP_APPOINTMENT_RECOMMENDED","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
# SAMPLE_COLLECTION
add_row("SAMPLE_COLLECTION","SAMPLE_COLLECTION_DECISION","DECISION","ATTACHED","immer","Nein","Nein","Ja","Nein")
add_row("SAMPLE_COLLECTION","URINE_SAMPLE_INSTRUCTIONS","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("SAMPLE_COLLECTION","STOOL_SAMPLE_INSTRUCTIONS","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("SAMPLE_COLLECTION","SAMPLE_HANDOVER","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("SAMPLE_COLLECTION","URINE_SAMPLE_ONSITE","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("SAMPLE_COLLECTION","LAB_SAMPLE_FOLLOWUP_APPOINTMENT_RECOMMENDED","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("SAMPLE_COLLECTION","LAB_RESULT_TIME","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
# ONBOARDING
add_row("ONBOARDING","ONBOARDING_IDENTITY_CLARIFICATION_REQUIRED","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("ONBOARDING","ONBOARDING_PROVIDE_IDENTITY_DATA","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("ONBOARDING","ONBOARDING_DATA_MISSING_CONTEXT","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("ONBOARDING","ONBOARDING_WRONG_PRACTICE_NOTICE","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("ONBOARDING","DOCUMENT_UPLOAD","ACTION","ATTACHED","config-abhängig","Ja","Ja","Nein","Nein")
add_row("ONBOARDING","INSURANCE_DATA_APP_TRANSFER","ACTION","SHARED_BOTTOM","bedingt","Ja","Ja","Nein","Nein")
add_row("ONBOARDING","ONBOARDING_DOCTOLIB_INFO","EXPLANATION","ATTACHED","config-abhängig","Nein","Nein","Nein","Nein")
# BILLING
add_row("BILLING","BILLING_COST_NOT_COVERED","EXPLANATION","ATTACHED","bedingt","Ja","Nein","Nein","Nein")
add_row("BILLING","BILLING_NOT_COVERED_BY_STATUTORY","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("BILLING","BILLING_GOA_BILLING","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("BILLING","BILLING_ONSITE_PAYMENT","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("BILLING","BILLING_CONTACT_EXTERNAL_PARTY","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("BILLING","BILLING_ADDRESS_UPDATE_REQUESTED","ACTION","ATTACHED","bedingt","Nein","Ja","Nein","Nein")
add_row("BILLING","INSURANCE_DATA_APP_TRANSFER","ACTION","SHARED_BOTTOM","bedingt","Ja","Ja","Nein","Nein")
add_row("BILLING","BILLING_INVOICE_TIMING","EXPLANATION","ATTACHED","config-abhängig","Nein","Nein","Nein","Nein")
# TECH_SUPPORT
add_row("TECH_SUPPORT","TECH_VIDEO_NOT_WORKING","EXPLANATION","ATTACHED","config-abhängig","Nein","Nein","Nein","Nein")

def build_matrix(wb):
    ws = wb.create_sheet("Profile_Checkpoint_Matrix")
    headers = [
        "Profile ID","Checkpoint ID","Typ","Placement",
        "Trigger/Bedingung","Shared?","Action?","Decision?","Intro?"
    ]
    ws.append(headers)
    for row in MATRIX_ROWS:
        ws.append(list(row))
    style_header_row(ws)
    style_data_rows(ws)
    auto_width(ws)
    freeze_and_filter(ws)

# ---------------------------------------------------------------------------
# Sheet 4: Action_Checkpoints
# ---------------------------------------------------------------------------
ACTION_CHECKPOINTS = [
    # (id, trigger, placement, profiles, description)
    ("DIGITAL_REQUEST","bedingt (availableActionIds)","SHARED_BOTTOM","alle außer ACUTE_CARE","Digitale Anfrage stellen"),
    ("BOOK_APPOINTMENT","bedingt","SHARED_BOTTOM","APPOINTMENT, MEDICAL_DOCUMENTS, HEILMITTELVERORDNUNG","Termin buchen"),
    ("INSURANCE_DATA_APP_TRANSFER","bedingt","SHARED_BOTTOM","AU, PRESCRIPTION, REFERRAL, APPOINTMENT, ONBOARDING, BILLING","Versicherungsdaten übertragen"),
    ("CONTROL_APPOINTMENT_RECOMMENDED","bedingt","SHARED_BOTTOM","AU, PRESCRIPTION, HOSPITAL_ADMISSION","Kontrolltermin empfohlen"),
    ("ACUTE_OPEN_CONSULTATION_ACTION","config-abhängig (openConsultationDays…)","ATTACHED","AU, ACUTE_CARE, APPOINTMENT","Offene Sprechstunde buchen"),
    ("CARE_CHANNEL_CHOICE","bedingt","ATTACHED","AU, ACUTE_CARE","Kanal-Auswahl"),
    ("DOCUMENT_UPLOAD","config-abhängig (uploadPlatformName…)","ATTACHED","PRESCRIPTION, HEILMITTELVERORDNUNG, HOSPITAL_ADMISSION, APPOINTMENT, MEDICAL_DOCUMENTS, ONBOARDING","Dokument hochladen"),
    ("E_RECIPE_USE","bedingt","ATTACHED","PRESCRIPTION","E-Rezept verwenden"),
    ("PHARMACY_INFORMATION","bedingt","ATTACHED","PRESCRIPTION","Apothekeninfo"),
    ("DIGITAL_REQUEST_PROCESSING_TIME","config-abhängig (digitalRequestProcessingTimeMin…)","ATTACHED","AU","Bearbeitungszeit digitale Anfrage"),
    ("REF_BOOKING_CODE_PROCESS","bedingt","ATTACHED","REFERRAL","Buchungscode-Prozess"),
    ("REF_ORIGINAL_VS_PDF","immer aktiv","ATTACHED","REFERRAL","Original vs. PDF Überweisung"),
    ("TRANSPORT_QUESTIONNAIRE_REQUEST","immer aktiv","ATTACHED","HOSPITAL_ADMISSION","Transportmittel-Fragebogen"),
    ("IMMUNIZATION_BOOK_VACCINATION","bedingt","ATTACHED","IMMUNIZATION","Impftermin buchen"),
    ("IMMUNIZATION_BOOK_COUNSELING","bedingt","ATTACHED","IMMUNIZATION","Impfberatung buchen"),
    ("IMMUNIZATION_BRING_VACCINATION_RECORD","standard aktiv; suppressed bei TRAVEL_MEDICINE","ATTACHED","IMMUNIZATION","Impfausweis mitbringen"),
    ("LAB_APPOINTMENT_INTERNAL","config-abhängig (doctorOrderBookingCode)","ATTACHED","LAB","Interner Labortermin"),
    ("LAB_APPOINTMENT_EXTERNAL","bedingt","ATTACHED","LAB","Externer Labortermin"),
    ("LAB_BRING_REFERRAL","bedingt","ATTACHED","LAB","Überweisungsschein mitbringen"),
    ("LAB_COST_COVERED_BY_REFERRAL","bedingt","ATTACHED","LAB","Kosten per Überweisung gedeckt"),
    ("LAB_SELF_PAYER_NOTE","bedingt","ATTACHED","LAB","Selbstzahler-Hinweis"),
    ("LAB_FASTING_REQUIRED","bedingt; suppressed bei MPU","ATTACHED","LAB","Nüchternheit erforderlich"),
    ("LAB_RESULT_TIME","bedingt","ATTACHED","LAB, SAMPLE_COLLECTION","Ergebniszeit"),
    ("LAB_SAMPLE_FOLLOWUP_APPOINTMENT_RECOMMENDED","bedingt","ATTACHED","LAB, SAMPLE_COLLECTION","Folgetermin Probe"),
    ("URINE_SAMPLE_INSTRUCTIONS","bedingt","ATTACHED","SAMPLE_COLLECTION","Anleitung Urinprobe"),
    ("STOOL_SAMPLE_INSTRUCTIONS","bedingt","ATTACHED","SAMPLE_COLLECTION","Anleitung Stuhlprobe"),
    ("SAMPLE_HANDOVER","bedingt","ATTACHED","SAMPLE_COLLECTION","Probenabgabe Info"),
    ("URINE_SAMPLE_ONSITE","bedingt","ATTACHED","SAMPLE_COLLECTION","Urinprobe vor Ort"),
    ("ONBOARDING_IDENTITY_CLARIFICATION_REQUIRED","bedingt","ATTACHED","ONBOARDING","Identitätsklärung nötig"),
    ("ONBOARDING_PROVIDE_IDENTITY_DATA","bedingt","ATTACHED","ONBOARDING","Identitätsdaten übermitteln"),
    ("ONBOARDING_DATA_MISSING_CONTEXT","bedingt","ATTACHED","ONBOARDING","Fehlende Daten Kontext"),
    ("ONBOARDING_WRONG_PRACTICE_NOTICE","bedingt","ATTACHED","ONBOARDING","Falsche Praxis Hinweis"),
    ("BILLING_NOT_COVERED_BY_STATUTORY","bedingt","ATTACHED","BILLING","Nicht von GKV gedeckt"),
    ("BILLING_GOA_BILLING","bedingt","ATTACHED","BILLING","GOÄ-Abrechnung"),
    ("BILLING_ONSITE_PAYMENT","bedingt","ATTACHED","BILLING","Zahlung vor Ort"),
    ("BILLING_CONTACT_EXTERNAL_PARTY","bedingt","ATTACHED","BILLING","Externer Ansprechpartner"),
    ("BILLING_ADDRESS_UPDATE_REQUESTED","bedingt","ATTACHED","BILLING","Adressänderung angefordert"),
    ("APPOINTMENT_BOOK_CHECKUP","bedingt","ATTACHED","APPOINTMENT","Check-up Termin buchen"),
    ("APPOINTMENT_BOOK_FINDINGS_REVIEW","config-abhängig (findingsReviewBookingCode)","ATTACHED","APPOINTMENT","Befundbesprechung buchen"),
    ("APPOINTMENT_BOOK_CHECKUP_SECOND","config-abhängig (checkupSecondBookingCode)","ATTACHED","APPOINTMENT","Check-up 2. Termin buchen"),
    ("APPOINTMENT_BOOK_CHRONIC_CONTROL","config-abhängig (chronicControlBookingCode)","ATTACHED","APPOINTMENT","Chronikerkontrolle buchen"),
    ("APPOINTMENT_BOOK_EKG_ORDER","config-abhängig (doctorOrderBookingCode)","ATTACHED","APPOINTMENT","EKG-Termin buchen"),
    ("APPOINTMENT_INFO_CHECKUP","bedingt","ATTACHED","APPOINTMENT","Info Check-up"),
    ("APPOINTMENT_INFO_CHRONIC_CONTROL","bedingt","ATTACHED","APPOINTMENT","Info Chronikertermin"),
    ("APPOINTMENT_INFO_FINDINGS_REVIEW","bedingt","ATTACHED","APPOINTMENT","Info Befundbesprechung"),
    ("APPOINTMENT_INFO_EKG_ORDER","bedingt","ATTACHED","APPOINTMENT","Info EKG-Termin"),
    ("APPOINTMENT_INFO_CHECKUP_SECOND","bedingt","ATTACHED","APPOINTMENT","Info Check-up 2. Termin"),
    ("APPOINTMENT_INFO_PSYCHOSOMATIC","bedingt","ATTACHED","APPOINTMENT","Info Psychosomatik"),
    ("APPOINTMENT_INFO_PHYSICAL_MEDICINE","bedingt","ATTACHED","APPOINTMENT","Info Physikalische Medizin"),
    ("APPOINTMENT_INFO_SURGERY","bedingt","ATTACHED","APPOINTMENT","Info Chirurgie"),
    ("APPOINTMENT_INFO_SPECIALTY_VISIT","bedingt","ATTACHED","APPOINTMENT","Info Spezialsprechstunde"),
    ("AU_NEW_PATIENT_3DAY_LIMIT","bedingt","ATTACHED","AU","3-Tage-Limit Neupatient AU"),
    ("PAYMENT_ONSITE_INFO","bedingt","ATTACHED","MEDICAL_DOCUMENTS","Zahlung vor Ort Info"),
]

def build_action_checkpoints(wb):
    ws = wb.create_sheet("Action_Checkpoints")
    headers = ["Checkpoint ID","Trigger","Placement","In welchen Profilen verwendet","Kurzbeschreibung"]
    ws.append(headers)
    for row in ACTION_CHECKPOINTS:
        ws.append(list(row))
    style_header_row(ws)
    style_data_rows(ws)
    auto_width(ws)
    freeze_and_filter(ws)

# ---------------------------------------------------------------------------
# Sheet 5: Config_Dependent
# ---------------------------------------------------------------------------
CONFIG_DEP = [
    ("ACUTE_OPEN_CONSULTATION_ACTION","openConsultationDays, openConsultationHours, openConsultationCapacityLimited","AU, ACUTE_CARE, APPOINTMENT","ACTION"),
    ("DIGITAL_REQUEST_PROCESSING_TIME","digitalRequestProcessingTimeMin, digitalRequestProcessingTimeMax, digitalRequestProcessingTimeUnit","AU","ACTION"),
    ("LAB_APPOINTMENT_INTERNAL","doctorOrderBookingCode","LAB","ACTION"),
    ("APPOINTMENT_BOOK_FINDINGS_REVIEW","bookingCalendarName, findingsReviewBookingCode","APPOINTMENT","ACTION"),
    ("APPOINTMENT_BOOK_CHECKUP_SECOND","bookingCalendarName, checkupSecondBookingCode","APPOINTMENT","ACTION"),
    ("APPOINTMENT_BOOK_CHRONIC_CONTROL","bookingCalendarName, chronicControlBookingCode","APPOINTMENT","ACTION"),
    ("APPOINTMENT_BOOK_EKG_ORDER","bookingCalendarName, doctorOrderBookingCode","APPOINTMENT","ACTION"),
    ("DOCUMENT_UPLOAD","uploadPlatformName, uploadPlatformAccountLabel","PRESCRIPTION, HEILMITTELVERORDNUNG, HOSPITAL_ADMISSION, APPOINTMENT, MEDICAL_DOCUMENTS, ONBOARDING","ACTION"),
    ("ONBOARDING_DOCTOLIB_INFO","uploadPlatformName, uploadPlatformAccountLabel","ONBOARDING","EXPLANATION"),
    ("BILLING_INVOICE_TIMING","billingCycleLabel","BILLING","EXPLANATION"),
    ("TECH_VIDEO_NOT_WORKING","videoSupportContact","TECH_SUPPORT","EXPLANATION"),
]

def build_config_dependent(wb):
    ws = wb.create_sheet("Config_Dependent")
    headers = ["Checkpoint ID","verwendete _cfg Felder","Profil(e)","Typ"]
    ws.append(headers)
    for row in CONFIG_DEP:
        ws.append(list(row))
    style_header_row(ws)
    style_data_rows(ws)
    auto_width(ws)
    freeze_and_filter(ws)

# ---------------------------------------------------------------------------
# Sheet 6: Reuse_Overview
# ---------------------------------------------------------------------------
REUSE = [
    ("DIGITAL_REQUEST", 13, "alle außer ACUTE_CARE", "GLOBAL"),
    ("BOOK_APPOINTMENT", 13, "alle außer ONBOARDING", "GLOBAL"),
    ("INSURANCE_DATA_APP_TRANSFER", 6, "AU, PRESCRIPTION, REFERRAL, APPOINTMENT, ONBOARDING, BILLING", "GLOBAL"),
    ("DOCUMENT_UPLOAD", 6, "PRESCRIPTION, HEILMITTELVERORDNUNG, HOSPITAL_ADMISSION, APPOINTMENT, MEDICAL_DOCUMENTS, ONBOARDING", "GLOBAL (config-abhängig)"),
    ("ACUTE_OPEN_CONSULTATION_ACTION", 3, "AU, ACUTE_CARE, APPOINTMENT", "GLOBAL"),
    ("CONTROL_APPOINTMENT_RECOMMENDED", 3, "AU, PRESCRIPTION, HOSPITAL_ADMISSION", "GLOBAL"),
    ("CONTACT_PERSON_INFO", 4, "AU, REFERRAL, PRESCRIPTION, ONBOARDING", "GLOBAL"),
    ("CARE_CHANNEL_CHOICE", 2, "AU, ACUTE_CARE", "GLOBAL"),
    ("BILLING_COST_NOT_COVERED", 2, "LAB, BILLING", "SPECIFIC/SHARED"),
    ("APPOINTMENT_DATA_INCOMPLETE", 2, "LAB, APPOINTMENT", "SPECIFIC/SHARED"),
    ("LAB_SAMPLE_FOLLOWUP_APPOINTMENT_RECOMMENDED", 2, "LAB, SAMPLE_COLLECTION", "SPECIFIC"),
    ("LAB_RESULT_TIME", 2, "LAB, SAMPLE_COLLECTION", "SPECIFIC"),
    ("PAYMENT_ONSITE_INFO", 1, "MEDICAL_DOCUMENTS", "GLOBAL"),
    ("TRANSPORT_QUESTIONNAIRE_REQUEST", 1, "HOSPITAL_ADMISSION", "GLOBAL"),
]

def build_reuse(wb):
    ws = wb.create_sheet("Reuse_Overview")
    headers = ["Checkpoint ID","Anzahl Profile","Verwendet in","Shared/global"]
    ws.append(headers)
    for row in REUSE:
        ws.append(list(row))
    style_header_row(ws)
    style_data_rows(ws)
    auto_width(ws)
    freeze_and_filter(ws)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_profiles(wb)
    build_checkpoints(wb)
    build_matrix(wb)
    build_action_checkpoints(wb)
    build_config_dependent(wb)
    build_reuse(wb)

    wb.save(OUT)
    print(f"Gespeichert: {OUT}")
    print(f"Sheets: {wb.sheetnames}")

if __name__ == "__main__":
    main()
